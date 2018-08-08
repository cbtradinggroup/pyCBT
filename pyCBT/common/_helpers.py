"""_helpers module

This module contains useful functions for the pyCBT scripts.
"""

# TODO: implement naming the instrument
# TODO: implement argument for naming the calendars
# TODO: implement several values-for-one-option parsing
# TODO: add interactive option
# TODO: add verbose option
def parse_args(*args, **kwargs):
    """Returns the passed command line arguments.
    """
    import argparse, sys, pytz
    from datetime import datetime

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "url",
        help="The instrument url in www.investing.com. E.g., www.investing.com/equities/apple-computer-inc."
    )
    parser.add_argument(
    	"--resolution",
    	choices=["Daily", "Weekly", "Monthly"],
    	default="Daily",
    	help="The time resolution of data"
    )
    parser.add_argument(
        "--from-date",
        help="Start date in dataset. Defaults to first register in table.",
        default=datetime(2000, 1, 1).strftime("%Y-%m-%d")
    )
    parser.add_argument(
        "--to-date",
        help="End date in dataset. Defaults to now in New York.",
        default=datetime.now(tz=pytz.timezone("America/New_York")).strftime("%Y-%m-%d")
    )
    parser.add_argument(
        "--save-to", "-s",
        help="Stores the dataset in the given filename. Supported extensions are .csv (default if extension is missing) or .xlsx file. If .xlsx and a second argument (comma-separated) value is given, it is taken to be the name of the sheet."
    )
    args = parser.parse_args()

    return args._get_args(), dict(args._get_kwargs())

def dump_data(*args, **kwargs):
    """Writes data in screen or into a file.
    """
    import os, sys
    from pyCBT.common.files import exist
    from openpyxl import load_workbook

    dataframe, = args
    if kwargs.get("save_to"):
        if "," in kwargs["save_to"]: filename, sheetname = kwargs["save_to"].split(",")
        else: filename, sheetname = kwargs.get("save_to"), "sheet_001"
        if filename.endswith(".xlsx") and exist(filename):
            # TODO: if the file exist, ask the user
            book = load_workbook(filename)
            with pd.ExcelWriter(filename, engine="openpyxl") as excel_writer:
                excel_writer.book = book
                if sheetname in book.sheetnames:
                    match = re.match("(\w+)_(\d+)", sheetname)
                    if not match: sheetname += "_{0:03d}"
                    else: sheetname = string.join([match.groups()[0], "{0:03d}"], "_")
                i = 1
                while sheetname.format(i) in book.sheetnames: i += 1
                sheetname = sheetname.format(i)
                dataframe.to_excel(excel_writer, sheet_name=sheetname)
                excel_writer.save()
            book.close()
        elif filename.endswith(".xlsx"):
            with pd.ExcelWriter(filename, engine="openpyxl") as excel_writer:
                dataframe.to_excel(excel_writer, sheet_name=sheetname)
                excel_writer.save()
        else:
            # TODO: if the file exist, ask the user
            if filename.split(".")[-1] != "csv": filename += ".csv"
            dataframe.reset_index().to_csv(filename, index=False, line_terminator=os.linesep)
    else:
        dataframe.reset_index().to_csv(sys.stdout, index=False, line_terminator=os.linesep)

    return None
